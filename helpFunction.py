import openpyxl
import datetime
import os
import math
today = datetime.date.today()
oneday = datetime.timedelta(days=1)
yesterday = str(today - oneday)
month = str(datetime.datetime.now().month) + "月"
save_dir = "save/"+str(today) + '/'
save_name = save_dir + str(today)#缺后缀，在用到的时候，要拼接
def filePathOfEachCompanyList(n):
    return save_name + "-线程" +str(n) + '要跑的机构名单.txt'
file_path_of_tmp=save_name + "-tmp.txt"
url_save_txt = filePathOfEachCompanyList(1)

#分N份保存到N个txt
#前面的机构，持股数量多达几千，后面的机构少到只有1个，所以用类似斗地主发牌方式，目的：每个线程分的均匀一点
def saveAndSplit(source,slice_num): # source=[[机构名1，url1，持有机构数量],[机构名2，url2，持有机构数量],[]]
    total=len(source)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    for i in range(slice_num):
        with open(filePathOfEachCompanyList(i+1), 'w', encoding='utf-8') as f:
            for j in range(i, total,slice_num):
                line= '-|-'.join(str(k) for k in source[j])+"\n" # 原本用的“,”，但是有的机构名称本身含有逗号，引起错位，改用其他符号
                f.write(line)
def saveResult(all_result,n):#n是第几个线程
    excel_name=save_name+"-线程"+str(n)+".xlsx"
    if not os.path.exists(excel_name):
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(excel_name)
    if "all" in wb.sheetnames:
        ws = wb["all"]
    else:
        ws = wb.create_sheet('all',0)
    for each in all_result:
        ws.append(each)
    wb.save(excel_name)
def saveTmp(list):#[机构名1，url1，持有机构数量],把那些不成功的机构写入tmp。txt
    with open(file_path_of_tmp, 'a',encoding='utf-8') as f:# a 是续写！ w 会覆盖重写！
        line = '-|-'.join(str(k) for k in list) + "\n"  # 原本用的“,”，但是有的机构名称本身含有逗号，引起错位，改用其他符号
        f.write(line)
def read_url(file_path): # 读取分好的的urls
    with open(file_path, 'r', encoding='utf-8') as f:
        urls = f.readlines()
    return urls # →['机构名1,url1,持有机构数量\n', '机构名1,url1,持有机构数量\n']
def getUrl(file_path):#从.txt取机构名称和url
    urls=read_url(file_path)
    result=[]
    for line in urls:
        l=line.strip().split("-|-")
        result.append(l)
    return result # →[[机构名，url，持有机构数量],[]]
def log(string,n):#记录每个线程的log输出
    with open(save_name + "-log-" + str(n) + '.txt', 'a', encoding='utf-8') as f:
        line = string + "\n"
        f.write(line)





